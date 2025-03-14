import pandas as pd
import pulp as pl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import re

# 设置工作目录并读取数据
raw_data = pd.read_excel("input.xlsx", sheet_name=0)
quota_limit = pd.read_excel("input.xlsx", sheet_name=1)

# 数据聚合处理
if len(raw_data.columns) > 1:
    data = raw_data.groupby(raw_data.columns.tolist()[1:]).size().reset_index(name='raw')
else:
    data = raw_data.iloc[:, 0].value_counts().reset_index()
    data.columns = [raw_data.columns[0], 'raw']

# 创建变量名
data['x'] = ['x' + str(i) for i in range(1, len(data) + 1)]


# 约束条件转换函数（关键修复部分）
def process_condition(cond):
    # 转换data$x为全选条件
    if cond.strip() == "data$x":
        return "index >= 0"  # 选择所有行

    # 移除data$x前缀
    cond = re.sub(r"data\$x\[(.*)\]", r"\1", cond)

    # 转换列引用格式
    cond = re.sub(r"data\$(\w+)", r"\1", cond)

    # 转换多值条件
    cond = re.sub(r"(\w+==)(\d+)\s*\|",
                  lambda m: f"({m.group(1)}{m.group(2)}) | ", cond)

    # 转换字符串条件
    cond = re.sub(r"(\w+)=='(\w+)'", r"\1 == '\2'", cond)

    # 转换数值范围条件
    cond = re.sub(r"\((\w+)==(\d+)\s*(\|.*){4,}\)",
                  lambda m: f"{m.group(1)}.isin([{','.join(m.group(0).split('|'))}])", cond)

    return cond


# 构建约束矩阵
n = len(quota_limit)
constraints = []
for i in range(n):
    original_cond = quota_limit.iloc[i, 0]
    cond = process_condition(original_cond)

    # 特殊处理全选条件
    if "index >= 0" in cond:
        selected = data
    else:
        try:
            selected = data.query(cond)
        except:
            print(f"Error in condition: {cond}")
            raise

    constraints.append(selected['x'].tolist())

# 创建线性规划问题
prob = pl.LpProblem("Selection_Problem", pl.LpMaximize)

# 创建决策变量
x_vars = {row.x: pl.LpVariable(row.x, lowBound=0, cat='Integer') for _, row in data.iterrows()}

# 目标函数
prob += pl.lpSum(x_vars.values())

# 添加约束（修复上下限处理）
for i in range(n):
    cons_vars = [x_vars[x] for x in constraints[i]]

    # 上限约束
    if not pd.isna(quota_limit.iloc[i, 2]):
        prob += pl.lpSum(cons_vars) <= quota_limit.iloc[i, 2]

    # 下限约束
    if not pd.isna(quota_limit.iloc[i, 1]):
        prob += pl.lpSum(cons_vars) >= quota_limit.iloc[i, 1]

# 求解问题
prob.solve()



# 处理结果
data['solution'] = [int(pl.value(x_vars[x])) for x in data['x']]

# 生成组合键
if len(data.columns) > 3:
    data['combine'] = data.iloc[:, :-2].astype(str).apply(lambda x: ''.join(x), axis=1)
else:
    data['combine'] = data.iloc[:, 0].astype(str)


# 生成最终选择结果
def cf(array):
    return array.groupby(array).cumcount() + 1


raw_data['combine'] = raw_data.astype(str).apply(lambda x: ''.join(x), axis=1)
raw_data['accum'] = cf(raw_data['combine'])
merged = pd.merge(raw_data, data[['combine', 'solution']], on='combine')
merged['select'] = (merged['accum'] <= merged['solution']).astype(int)

# 结果对比分析
compare = quota_limit.copy()
compare['raw'] = 0
compare['raw_gap'] = 0
compare['plan'] = 0
compare['plan_gap'] = 0

for i in range(n):
    cond = process_condition(quota_limit.iloc[i, 0])
    compare.iloc[i, 2] = len(data.query(cond))
    compare.iloc[i, 4] = data.query(cond)['solution'].sum()

compare['raw_gap'] = compare['raw'] - compare['max']
compare['plan_gap'] = compare['plan'] - compare['max']

# 创建Excel输出
wb = Workbook()
ws1 = wb.active
ws1.title = "samples_select"
ws2 = wb.create_sheet("result")
ws3 = wb.create_sheet("compare")

# 写入数据
for r in dataframe_to_rows(merged.drop(['combine', 'accum', 'solution'], axis=1), index=False, header=True):
    ws1.append(r)

for r in dataframe_to_rows(data.drop(['x', 'combine'], axis=1), index=False, header=True):
    ws2.append(r)

for r in dataframe_to_rows(compare, index=False, header=True):
    ws3.append(r)

# 设置样式
header_fill = PatternFill(start_color="CD8500", end_color="CD8500", fill_type="solid")
center_aligned = Alignment(horizontal="center")

for ws in [ws1, ws2, ws3]:
    for cell in ws[1]:
        cell.fill = header_fill
        cell.alignment = center_aligned
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = center_aligned

# 特殊处理第一个全量约束（data$x）
if len(constraints[0]) == len(data):
    print("全局约束处理成功")
else:
    print("警告：全局约束可能未正确应用")

# 保存文件前强制刷新数据
wb.save("output.xlsx")
print("输出文件已保存")

# 输出结果验证
if compare['plan'].sum() > 0:
    print("Success!")
else:
    print("Fail~")

