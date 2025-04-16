import pandas as pd
import numpy as np
import os
import logging
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from pulp import *

file_name = 'input.xlsx'
TIMEOUT = 100
# 配置日志
script_dir = os.path.dirname(os.path.abspath(__file__))
log_dir = os.path.join(script_dir, 'logs')
os.makedirs(log_dir, exist_ok=True)
log_file = os.path.join(log_dir, f'sampling_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file, encoding='utf-8'),
        logging.StreamHandler()
    ]
)

def convert_r_condition(condition):
    """
    将R语言格式转换为pandas的格式
    """
    if condition == 'total':
        return condition
    
    if not condition.startswith('data$'):
        return condition
    
    # 检查是否是R格式
    if condition.startswith('data$x'):
        if condition == 'data$x':
            return 'total'       
        # 提取括号内的条件
        match = re.search(r'\[(.*?)\]', condition)
        if match:
            r_cond = match.group(1)
            
            # 处理%in%操作符
            in_match = re.search(r'(\w+)\s*%in%\s*c\((.*?)\)', r_cond)
            if in_match:
                col = in_match.group(1)
                values = in_match.group(2)
                value_list = [v.strip() for v in values.split(',')]
                for i, v in enumerate(value_list):
                    try:
                        value_list[i] = int(v)
                    except ValueError:
                        try:
                            value_list[i] = float(v)
                        except ValueError:
                            value_list[i] = f"'{v}'"
                return f"{col}.isin({value_list})"
            
            # 替换data$为空
            r_cond = r_cond.replace('data$', '')
            # 将 )|( 替换为 ) | (
            if ')|(' in r_cond:
                r_cond = r_cond.replace(')|(', ') | (')
            
            # 处理括号
            # 外层括号
            if r_cond.startswith('(') and r_cond.endswith(')'):
                # 嵌套括号
                if r_cond.startswith('((') and r_cond.endswith('))'):
                    # 保留一层括号
                    r_cond = r_cond[1:-1]
            
            # 确保|周围有空格
            r_cond = r_cond.replace('|', ' | ')
            r_cond = re.sub(r'\s+\|\s+', ' | ', r_cond)
            
            # 确保&周围有空格
            r_cond = r_cond.replace('&', ' & ')
            r_cond = re.sub(r'\s+&\s+', ' & ', r_cond)
            
            return r_cond
    
    return condition

def validate_input_data(data_df, cond_df):
    """验证输入数据"""
    # 检查数据
    if data_df.empty:
        raise ValueError("数据表为空，请检查输入文件")
    if cond_df.empty:
        raise ValueError("条件表为空，请检查输入文件")
    
    # 检查条件表必要列
    required_columns = ['condition', 'quota', 'min', 'max']
    missing_columns = [col for col in required_columns if col not in cond_df.columns]
    if missing_columns:
        raise ValueError(f"条件表缺少列: {', '.join(missing_columns)}")
    
    # 转换R格式条件,更新条件表
    for idx, row in cond_df.iterrows():
        cond_df.at[idx, 'condition'] = convert_r_condition(row['condition'])
    
    # 检查total条件是否存在
    if 'total' not in cond_df['condition'].values:
        raise ValueError("条件表中缺少'total'条件")
    
    # 验证条件语法
    for _, row in cond_df[cond_df['condition'] != 'total'].iterrows():
        try:
            if not isinstance(row['condition'], str):
                raise ValueError(f"条件必须是字符串: {row['condition']}")
            data_df.eval(row['condition'])
        except Exception as e:
            raise ValueError(f"条件语法错误: {row['condition']}，错误信息: {str(e)}")
    
    logging.info(f"输入数据验证通过: {len(data_df)}行数据, {len(cond_df)}个条件")
    return True

def parse_conditions(cond_df):
    """解析抽数条件"""
    conditions = []
    total_row = cond_df[cond_df['condition'] == 'total']
    if total_row.empty:
        raise ValueError("未找到total条件行")
    
    total = total_row['quota'].values[0]
    logging.info(f"总样本数量: {total}")
    
    for _, row in cond_df[cond_df['condition'] != 'total'].iterrows():
        # 验证max和min的有效性
        min_val = max(0, row['min']) if pd.notna(row['min']) else 0
        max_val = max(min_val, row['max']) if pd.notna(row['max']) else float('inf')
        
        conditions.append({
            'condition': row['condition'],
            'quota': row['quota'],
            'min': min_val,
            'max': max_val
        })
        logging.info(f"条件: {row['condition']}, 配额: {row['quota']}, 最小: {min_val}, 最大: {max_val}")
    
    return total, conditions

def linear_programming_sampling(data_df, conditions, total, timeout=TIMEOUT, fallback=True):
    """使用线性规划进行抽数，使抽样结果尽量接近期望值
    
    参数:
        data_df: 数据DataFrame
        conditions: 条件
        total: 目标总样本量
        timeout: 求解超时时间(秒)
        fallback: 是否在复杂求解失败时回退到简单方法
    """
    logging.info(f"开始抽数，目标样本量: {total}")
    
    # 原始数据
    original_df = data_df.copy()
    
    # 创建组合标识符
    if data_df.shape[1] > 1:
        data_df['_combine_id'] = data_df.astype(str).apply(lambda x: ''.join(x), axis=1)
    else:
        data_df['_combine_id'] = data_df.iloc[:, 0].astype(str)
    
    # 计算每个组合的出现次数
    group_counts = data_df['_combine_id'].value_counts().reset_index()
    group_counts.columns = ['_combine_id', 'raw_count']
    
    # 分配变量名
    group_counts['var_name'] = [f'x{i+1}' for i in range(len(group_counts))]
    
    # 将原始数据映射到组变量
    id_to_var = dict(zip(group_counts['_combine_id'], group_counts['var_name']))
    data_df['_var_name'] = data_df['_combine_id'].map(id_to_var)
    
    # 判断条件数量，如果条件过多，使用优化的方法
    complex_conditions = len(conditions) > 15
    if complex_conditions:
        logging.info(f"检测到{len(conditions)}个条件，使用优化的求解方法")
    
    try:
        # 创建线性规划问题
        if complex_conditions:
            # 使用聚合偏差变量(所有条件偏差绝对值之和/条件数量)
            prob = LpProblem("Sampling_Problem", LpMinimize)
            
            # 创建决策变量
            var_dict = {}
            for _, row in group_counts.iterrows():
                var_name = row['var_name']
                raw_count = row['raw_count']
                var_dict[var_name] = LpVariable(var_name, lowBound=0, upBound=raw_count, cat='Integer')
            
            # 创建单个总偏差变量
            total_deviation = LpVariable("total_deviation", lowBound=0, cat='Continuous')
            
            # 目标函数：最小化总偏差
            prob += total_deviation, "Minimize_Total_Deviation"
            
            # 约束条件1：总样本量等于目标总量
            prob += lpSum(var_dict.values()) == total, "Total_Samples"
            
            # 约束条件2：计算每个条件的偏差并约束总偏差
            all_deviations = []
            for i, cond in enumerate(conditions):
                mask = data_df.eval(cond['condition'])
                cond_vars = data_df.loc[mask, '_var_name'].unique()
                
                if cond_vars.size > 0:
                    # 条件变量的总和表达式
                    cond_sum = lpSum([var_dict[var] for var in cond_vars if var in var_dict])
                    
                    # 各条件的偏差变量
                    deviation_var = LpVariable(f"dev_{i}", lowBound=0, cat='Continuous')
                    all_deviations.append(deviation_var)
                    
                    # 偏差约束
                    prob += deviation_var >= cond_sum - cond['quota'], f"Dev_Constraint_Pos_{i}"
                    prob += deviation_var >= cond['quota'] - cond_sum, f"Dev_Constraint_Neg_{i}"
                    
                    # 最小约束
                    prob += cond_sum >= cond['min'], f"Min_Constraint_{i}"
                    # 最大约束
                    if cond['max'] != float('inf'):
                        prob += cond_sum <= cond['max'], f"Max_Constraint_{i}"
            
            # 总偏差必须大于等于所有条件偏差的加权和
            prob += total_deviation >= lpSum(all_deviations) / len(conditions), "Total_Deviation_Constraint"
            
        else:
            # 原始版本：为每个条件使用一对偏差变量
            prob = LpProblem("Sampling_Problem", LpMinimize)
            
            # 创建决策变量
            var_dict = {}
            for _, row in group_counts.iterrows():
                var_name = row['var_name']
                raw_count = row['raw_count']
                var_dict[var_name] = LpVariable(var_name, lowBound=0, upBound=raw_count, cat='Integer')
            
            # 偏差变量 - 用于测量与配额的差异
            deviation_vars = {}
            for i, cond in enumerate(conditions):
                # 正偏差 (实际 > 配额)
                deviation_vars[f'pos_dev_{i}'] = LpVariable(f'pos_dev_{i}', lowBound=0, cat='Continuous')
                # 负偏差 (实际 < 配额)
                deviation_vars[f'neg_dev_{i}'] = LpVariable(f'neg_dev_{i}', lowBound=0, cat='Continuous')
            
            # 目标函数：最小化所有条件的偏差总和
            prob += lpSum([deviation_vars[f'pos_dev_{i}'] + deviation_vars[f'neg_dev_{i}'] for i in range(len(conditions))])
            
            # 约束条件1：总样本量等于目标总量
            prob += lpSum(var_dict.values()) == total, "Total_Samples"
            
            # 约束条件2：样本量在min和max之间，并设置偏差变量
            for i, cond in enumerate(conditions):
                mask = data_df.eval(cond['condition'])
                cond_vars = data_df.loc[mask, '_var_name'].unique()
                
                if cond_vars.size > 0:
                    # 创建条件变量的总和表达式
                    cond_sum = lpSum([var_dict[var] for var in cond_vars if var in var_dict])
                    
                    # 设置软约束 - 测量与配额的差异，但不强制精确等于
                    prob += cond_sum - cond['quota'] <= deviation_vars[f'pos_dev_{i}'], f"Pos_Deviation_Constraint_{i}"
                    prob += cond['quota'] - cond_sum <= deviation_vars[f'neg_dev_{i}'], f"Neg_Deviation_Constraint_{i}"
                    
                    # 最小约束
                    prob += cond_sum >= cond['min'], f"Min_Constraint_{i}"
                    # 最大约束
                    if cond['max'] != float('inf'):
                        prob += cond_sum <= cond['max'], f"Max_Constraint_{i}"
        
        # 约束条件3：每个组合的抽数不超过其原始数量
        for var_name, raw_count in zip(group_counts['var_name'], group_counts['raw_count']):
            prob += var_dict[var_name] <= raw_count, f"Raw_Limit_{var_name}"
        
        # 求解问题，设置超时
        logging.info(f"求解线性规划问题，超时时间: {timeout}秒...")
        solver = PULP_CBC_CMD(msg=False, timeLimit=timeout)
        prob.solve(solver)
        
        # 求解结果
        if LpStatus[prob.status] != 'Optimal':
            error_msg = f"线性规划求解失败，状态: {LpStatus[prob.status]}"
            logging.error(error_msg)
            if not fallback:
                raise ValueError(error_msg)
            else:
                raise Exception("尝试回退到简单方法")
        
        logging.info(f"线性规划求解成功，状态: {LpStatus[prob.status]}")
        
        # 分析偏差
        if complex_conditions:
            logging.info(f"总偏差: {prob.objective.value():.2f}")
        else:
            total_deviation = 0
            for i, cond in enumerate(conditions):
                pos_dev = prob.variablesDict()[f'pos_dev_{i}'].value()
                neg_dev = prob.variablesDict()[f'neg_dev_{i}'].value()
                actual_dev = pos_dev - neg_dev
                total_deviation += abs(actual_dev)
                logging.info(f"条件 '{cond['condition']}' 的偏差: {actual_dev:.2f} (目标: {cond['quota']})")
            
            logging.info(f"总偏差: {total_deviation:.2f}")
        
        # 提取结果
        results = {}
        for v in prob.variables():
            if v.name.startswith('x') and v.varValue > 0:
                results[v.name] = int(v.varValue)
        
    except Exception as e:
        if not fallback:
            raise ValueError(f"线性规划求解失败: {str(e)}")
        
        logging.warning(f"优化方法求解失败: {str(e)}，回退到简单方法")
        
        # 回退到简单方法
        prob = LpProblem("Sampling_Problem_Fallback", LpMaximize)
        
        # 创建决策变量
        var_dict = {}
        for _, row in group_counts.iterrows():
            var_name = row['var_name']
            raw_count = row['raw_count']
            var_dict[var_name] = LpVariable(var_name, lowBound=0, upBound=raw_count, cat='Integer')
        
        # 目标函数
        prob += lpSum(var_dict.values())
        
        # 约束条件1：总样本量等于total
        prob += lpSum(var_dict.values()) == total, "Total_Samples"
        
        # 约束条件2：样本量在min和max之间
        for i, cond in enumerate(conditions):
            mask = data_df.eval(cond['condition'])
            cond_vars = data_df.loc[mask, '_var_name'].unique()
            
            if cond_vars.size > 0:
                # 最小约束
                prob += lpSum([var_dict[var] for var in cond_vars if var in var_dict]) >= cond['min'], f"Min_Constraint_{i}"
                # 最大约束
                if cond['max'] != float('inf'):
                    prob += lpSum([var_dict[var] for var in cond_vars if var in var_dict]) <= cond['max'], f"Max_Constraint_{i}"
        
        # 约束条件3：每个组合的抽数不超过其原始数量
        for var_name, raw_count in zip(group_counts['var_name'], group_counts['raw_count']):
            prob += var_dict[var_name] <= raw_count, f"Raw_Limit_{var_name}"
        
        # 求解问题
        logging.info("使用简单方法求解线性规划问题...")
        prob.solve(PULP_CBC_CMD(msg=False))
        
        # 求解结果
        if LpStatus[prob.status] != 'Optimal':
            error_msg = f"简单方法线性规划求解失败，状态: {LpStatus[prob.status]}"
            logging.error(error_msg)
            raise ValueError(error_msg)
        
        logging.info(f"简单方法线性规划求解成功，状态: {LpStatus[prob.status]}")
        
        # 提取结果
        results = {}
        for v in prob.variables():
            if v.varValue > 0 and v.name.startswith('x'):
                results[v.name] = int(v.varValue)
    
    # 根据结果抽数
    sampled = pd.DataFrame()
    for var_name, sample_count in results.items():
        if sample_count > 0:
            # 属于该变量的所有行
            group_rows = data_df[data_df['_var_name'] == var_name]
            # 抽取的数量小于该组的总行数，随机抽取
            if sample_count < len(group_rows):
                group_sample = group_rows.sample(n=sample_count, random_state=42)
            else:
                group_sample = group_rows
            
            sampled = pd.concat([sampled, group_sample])
    
    # 移除临时列
    if '_combine_id' in sampled.columns:
        sampled = sampled.drop(['_combine_id', '_var_name'], axis=1)
    
    # 验证抽数结果
    if len(sampled) != total:
        error_msg = f"抽数失败：抽取的总样本量({len(sampled)})不等于目标总量({total})"
        logging.error(error_msg)
        raise ValueError(error_msg)
    
    # 验证各条件样本量是否在min到max之间
    for cond in conditions:
        condition_count = sampled.eval(cond['condition']).sum()
        if condition_count < cond['min']:
            error_msg = f"抽数失败：条件 '{cond['condition']}' 的抽取数量({condition_count})小于最小要求({cond['min']})"
            logging.error(error_msg)
            raise ValueError(error_msg)
        if cond['max'] != float('inf') and condition_count > cond['max']:
            error_msg = f"抽数失败：条件 '{cond['condition']}' 的抽取数量({condition_count})大于最大限制({cond['max']})"
            logging.error(error_msg)
            raise ValueError(error_msg)
    
    logging.info(f"抽数成功，共抽取 {len(sampled)} 个样本")
    
    # SL列
    original_df['SL'] = 0
    
    # 将抽取的样本标记为1
    sampled_indices = sampled.index
    original_df.loc[sampled_indices, 'SL'] = 1

    return sampled, original_df

def generate_report(sampled, conditions, original_data, total_quota):
    """生成报告"""
    report = []    
    
    total_row = {
        'condition': 'total',
        'quota': total_quota,
        'min': total_quota,
        'max': total_quota,
        'raw': len(original_data),
        'sampled': len(sampled),
        'gap': len(sampled) - total_quota,
    }
    report.append(total_row)
    
    # 添加各条件统计
    for cond in conditions:
        raw_count = original_data.eval(cond['condition']).sum()
        sampled_count = sampled.eval(cond['condition']).sum()
        
        # 验证min-max限制
        min_ok = sampled_count >= cond['min']
        max_ok = sampled_count <= cond['max'] if cond['max'] != float('inf') else True
        
        row = {
            'condition': cond['condition'],
            'quota': cond['quota'],
            'min': cond['min'],
            'max': cond['max'] if cond['max'] != float('inf') else 'N/A',
            'raw': raw_count,
            'sampled': sampled_count,
            'gap': sampled_count - cond['quota'],
        }
        report.append(row)
    
    return pd.DataFrame(report)

def format_excel(writer, report_df):
    """美化Excel"""
    workbook = writer.book
    worksheet = writer.sheets['报告']
    
    # 定义样式
    header_fill = PatternFill(start_color='FFCCCCFF', end_color='FFCCCCFF', fill_type='solid')
    header_font = Font(bold=True, color='FF000000')
    center_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # 应用样式
    for col_num, column in enumerate(report_df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # 调整列宽
    for col_num, column in enumerate(report_df.columns, 1):
        column_width = max(len(str(column)), report_df[column].astype(str).map(len).max())
        worksheet.column_dimensions[chr(64 + col_num)].width = column_width + 4
    
    # 应用样式
    for row_num in range(2, len(report_df) + 2):
        for col_num in range(1, len(report_df.columns) + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
    
    # 高亮显示
    for row_num, row in enumerate(report_df.itertuples(), 2):
        # 处理配额差距
        if hasattr(row, 'gap') and hasattr(row, 'quota') and row.quota > 0:
            gap_cell = worksheet.cell(row=row_num, column=report_df.columns.get_loc('gap') + 1)
            
            # 如果未达到配额，标黄色
            if getattr(row, 'gap') != 0:
                gap_cell.fill = PatternFill(start_color='FFFFCC00', end_color='FFFFCC00', fill_type='solid')
                gap_cell.font = Font(bold=True)
            else:
                gap_cell.fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')

def main():
    try:
        logging.info("开始抽数")
        
        script_dir = os.path.dirname(os.path.abspath(__file__))
        
        # 读取文件
        input_file = os.path.join(script_dir, file_name)
        if not os.path.exists(input_file):
            raise FileNotFoundError(f"输入文件 {input_file} 不存在")
            
        logging.info(f"读取输入文件: {input_file}")
        data_df = pd.read_excel(input_file, sheet_name=0)
        cond_df = pd.read_excel(input_file, sheet_name=1)
        
        # 验证输入
        validate_input_data(data_df, cond_df)
        
        # 解析条件
        total, conditions = parse_conditions(cond_df)
        
        # 执行线性规划抽数
        original_data = data_df.copy()
        try:
            # 设置超时时间，启用回退机制
            sampled, full_data = linear_programming_sampling(data_df, conditions, total, timeout=TIMEOUT, fallback=True)
            
            # 生成报告
            report_df = generate_report(sampled, conditions, original_data, total)
            
            # 保存结果
            output_file = os.path.join(script_dir, 'output.xlsx')
            logging.info(f"保存结果到: {output_file}")
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                full_data.to_excel(writer, sheet_name='样本', index=False)
                report_df.to_excel(writer, sheet_name='报告', index=False)
                format_excel(writer, report_df)
            
            logging.info(f"抽数成功，共抽取 {len(sampled)} 个样本，报告已保存到 {output_file}")
            print(f"\n抽数成功，共抽取 {len(sampled)} 个样本，报告已保存到 {output_file}")
            print(f"日志文件保存在: {log_file}")
            
        except ValueError as ve:
            # 抽数失败
            logging.error(f"抽数失败: {str(ve)}")
            print(f"\n抽数失败: {str(ve)}")
            print("请调整条件配置或数据后重试")
            print(f"详细错误信息请查看日志: {log_file}")
        
    except Exception as e:
        logging.error(f"程序执行出错: {str(e)}", exc_info=True)
        print(f"\n错误: {str(e)}")
        print(f"详细错误信息请查看日志: {log_file}")

if __name__ == '__main__':
    main()