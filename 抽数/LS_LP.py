import pandas as pd
import numpy as np
import os
import logging
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from pulp import *

# 文件名
file_name = 'input.xlsx'

# 配置日志
script_dir = os.path.dirname(os.path.abspath(__file__))
log_dir = os.path.join(script_dir, 'logs')
os.makedirs(log_dir, exist_ok=True)
log_file = os.path.join(log_dir, f'sampling_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file, encoding='utf-8'),
        logging.StreamHandler()
    ]
)

def convert_r_condition(condition):
    """
    将R语言格式的条件转换为pandas可识别的格式
    """
    if condition == 'total':
        return condition
    
    if not condition.startswith('data$'):
        return condition
    
    # 检查是否是R格式的条件
    if condition.startswith('data$x'):
        if condition == 'data$x':
            return 'total'       
        # 提取括号内的条件
        match = re.search(r'\[(.*?)\]', condition)
        if match:
            r_cond = match.group(1)
            
            # 处理%in%操作符
            in_match = re.search(r'(\w+)\s*%in%\s*c\((.*?)\)', r_cond)
            if in_match:
                col = in_match.group(1)
                values = in_match.group(2)
                value_list = [v.strip() for v in values.split(',')]
                for i, v in enumerate(value_list):
                    try:
                        value_list[i] = int(v)
                    except ValueError:
                        try:
                            value_list[i] = float(v)
                        except ValueError:
                            value_list[i] = f"'{v}'"
                return f"{col}.isin({value_list})"
            
            # 替换所有的data$为空
            r_cond = r_cond.replace('data$', '')
            
            # 特殊处理形如 (C104==1)|(C104==2) 的模式
            # 这种模式通常出现在多个OR条件的情况下
            if ')|(' in r_cond:
                # 将 )|( 替换为 ) | (
                r_cond = r_cond.replace(')|(', ') | (')
            
            # 处理括号
            # 首先检查是否有外层括号
            if r_cond.startswith('(') and r_cond.endswith(')'):
                # 检查是否有嵌套括号
                if r_cond.startswith('((') and r_cond.endswith('))'):
                    # 保留一层括号，去掉外层括号
                    r_cond = r_cond[1:-1]
                # 否则保持原样，不去掉括号
            
            # 处理OR条件 (|)
            # 确保|周围有空格
            r_cond = r_cond.replace('|', ' | ')
            r_cond = re.sub(r'\s+\|\s+', ' | ', r_cond)
            
            # 处理AND条件 (&)
            # 确保&周围有空格
            r_cond = r_cond.replace('&', ' & ')
            r_cond = re.sub(r'\s+&\s+', ' & ', r_cond)
            
            return r_cond
    
    return condition

def validate_input_data(data_df, cond_df):
    """验证输入数据"""
    # 检查数据
    if data_df.empty:
        raise ValueError("数据表为空，请检查输入文件")
    if cond_df.empty:
        raise ValueError("条件表为空，请检查输入文件")
    
    # 检查条件表必要列
    required_columns = ['condition', 'quota', 'min', 'max']
    missing_columns = [col for col in required_columns if col not in cond_df.columns]
    if missing_columns:
        raise ValueError(f"条件表缺少列: {', '.join(missing_columns)}")
    
    # 转换R格式条件并更新条件表
    for idx, row in cond_df.iterrows():
        cond_df.at[idx, 'condition'] = convert_r_condition(row['condition'])
    
    # 检查total条件是否存在
    if 'total' not in cond_df['condition'].values:
        raise ValueError("条件表中缺少'total'条件行")
    
    # 验证条件语法
    for _, row in cond_df[cond_df['condition'] != 'total'].iterrows():
        try:
            # 尝试评估条件
            if not isinstance(row['condition'], str):
                raise ValueError(f"条件必须是字符串: {row['condition']}")
            # 转换后的条件已经是pandas格式，可以直接eval
            data_df.eval(row['condition'])
        except Exception as e:
            raise ValueError(f"条件语法错误: {row['condition']}，错误信息: {str(e)}")
    
    logging.info(f"输入数据验证通过: {len(data_df)}行数据, {len(cond_df)}个条件")
    return True

def parse_conditions(cond_df):
    """解析抽数条件"""
    conditions = []
    total_row = cond_df[cond_df['condition'] == 'total']
    if total_row.empty:
        raise ValueError("未找到total条件行")
    
    total = total_row['quota'].values[0]
    logging.info(f"总抽样数量: {total}")
    
    for _, row in cond_df[cond_df['condition'] != 'total'].iterrows():
        # 确保min和max是有效的数值
        min_val = max(0, row['min']) if pd.notna(row['min']) else 0
        max_val = max(min_val, row['max']) if pd.notna(row['max']) else float('inf')
        
        # 条件已经在validate_input_data中转换过了，这里直接使用
        conditions.append({
            'condition': row['condition'],
            'quota': row['quota'],
            'min': min_val,
            'max': max_val
        })
        logging.info(f"条件: {row['condition']}, 配额: {row['quota']}, 最小: {min_val}, 最大: {max_val}")
    
    return total, conditions

def linear_programming_sampling(data_df, conditions, total):
    """使用线性规划进行抽样"""
    logging.info(f"开始抽样，目标总量: {total}")
    
    # 保存原始数据的副本，用于后续处理
    original_df = data_df.copy()
    
    # aggregate：为每个唯一的数据组合创建一个索引
    # 创建组合标识符
    if data_df.shape[1] > 1:
        data_df['_combine_id'] = data_df.astype(str).apply(lambda x: ''.join(x), axis=1)
    else:
        data_df['_combine_id'] = data_df.iloc[:, 0].astype(str)
    
    # 计算每个组合的出现次数
    group_counts = data_df['_combine_id'].value_counts().reset_index()
    group_counts.columns = ['_combine_id', 'raw_count']
    
    # 分配一个变量名
    group_counts['var_name'] = [f'x{i+1}' for i in range(len(group_counts))]
    
    # 将原始数据行映射到组变量
    id_to_var = dict(zip(group_counts['_combine_id'], group_counts['var_name']))
    data_df['_var_name'] = data_df['_combine_id'].map(id_to_var)
    
    # 创建线性规划
    prob = LpProblem("Sampling_Problem", LpMaximize)
    
    # 创建决策变量
    var_dict = {}
    for _, row in group_counts.iterrows():
        var_name = row['var_name']
        raw_count = row['raw_count']
        var_dict[var_name] = LpVariable(var_name, lowBound=0, upBound=raw_count, cat='Integer')
    
    # 目标函数
    prob += lpSum(var_dict.values())
    
    # 约束条件1：总抽样量等于目标总量
    prob += lpSum(var_dict.values()) == total, "Total_Samples"
    
    # 约束条件2：对每个条件，抽样量在min和max之间
    for i, cond in enumerate(conditions):
        # 找出满足条件的行
        mask = data_df.eval(cond['condition'])
        # 获取满足条件的变量名
        cond_vars = data_df.loc[mask, '_var_name'].unique()
        
        # 创建条件约束
        if cond_vars.size > 0:
            # 最小约束
            prob += lpSum([var_dict[var] for var in cond_vars if var in var_dict]) >= cond['min'], f"Min_Constraint_{i}"
            # 最大约束
            if cond['max'] != float('inf'):
                prob += lpSum([var_dict[var] for var in cond_vars if var in var_dict]) <= cond['max'], f"Max_Constraint_{i}"
    
    # 约束条件3：每个组合的抽样不超过其原始数量
    for var_name, raw_count in zip(group_counts['var_name'], group_counts['raw_count']):
        prob += var_dict[var_name] <= raw_count, f"Raw_Limit_{var_name}"
    
    # 求解问题
    logging.info("求解线性规划问题...")
    prob.solve(PULP_CBC_CMD(msg=False))
    
    # 检查求解状态
    if LpStatus[prob.status] != 'Optimal':
        error_msg = f"线性规划求解失败，状态: {LpStatus[prob.status]}"
        logging.error(error_msg)
        raise ValueError(error_msg)
    
    logging.info(f"线性规划求解成功，状态: {LpStatus[prob.status]}")
    
    # 提取结果
    results = {}
    for v in prob.variables():
        if v.varValue > 0:
            results[v.name] = int(v.varValue)
    
    # 根据结果抽样
    sampled = pd.DataFrame()
    for var_name, sample_count in results.items():
        if sample_count > 0:
            # 找出属于该变量的所有行
            group_rows = data_df[data_df['_var_name'] == var_name]
            # 如果需要抽取的数量小于该组的总行数，随机抽取
            if sample_count < len(group_rows):
                group_sample = group_rows.sample(n=sample_count, random_state=42)
            else:
                group_sample = group_rows
            
            sampled = pd.concat([sampled, group_sample])
    
    # 移除临时列
    if '_combine_id' in sampled.columns:
        sampled = sampled.drop(['_combine_id', '_var_name'], axis=1)
    
    # 验证抽样结果
    if len(sampled) != total:
        error_msg = f"抽样失败：抽取的总样本量({len(sampled)})不等于目标总量({total})"
        logging.error(error_msg)
        raise ValueError(error_msg)
    
    # 验证各条件的抽取数量是否在min到max之间
    for cond in conditions:
        condition_count = sampled.eval(cond['condition']).sum()
        if condition_count < cond['min']:
            error_msg = f"抽样失败：条件 '{cond['condition']}' 的抽取数量({condition_count})小于最小要求({cond['min']})"
            logging.error(error_msg)
            raise ValueError(error_msg)
        if cond['max'] != float('inf') and condition_count > cond['max']:
            error_msg = f"抽样失败：条件 '{cond['condition']}' 的抽取数量({condition_count})大于最大限制({cond['max']})"
            logging.error(error_msg)
            raise ValueError(error_msg)
    
    logging.info(f"抽样成功，共抽取 {len(sampled)} 个样本，所有条件都满足min-max限制")
    
    # 创建完整的输出数据集，包含所有原始数据，并添加SL列
    # 初始化SL列为0（未抽取）
    original_df['SL'] = 0
    
    # 将抽取的样本在SL列标记为1
    # 注意：需要根据索引匹配，因为sampled是data_df的子集
    sampled_indices = sampled.index
    original_df.loc[sampled_indices, 'SL'] = 1
    
    # 返回完整数据集（带SL标记）和抽样数据
    return sampled, original_df

def generate_report(sampled, conditions, original_data, total_quota):
    """生成详细的统计报告，包括min-max限制验证"""
    report = []
    
    # 添加总量统计
    # 使用传入的total_quota作为总抽样目标
    
    total_row = {
        'condition': 'total',
        'quota': total_quota,  # 使用传入的总抽样目标
        'min': total_quota,   # 总抽样目标即为最小值
        'max': total_quota,   # 总抽样目标即为最大值
        'raw': len(original_data),
        'raw_pct': 100.0,
        'sampled': len(sampled),
        'sampled_pct': 100.0,
        'gap': len(sampled) - total_quota,  # 使用正确的总抽样目标计算差距
        'completion': f"{len(sampled) / total_quota * 100:.1f}%" if total_quota > 0 else "N/A",  # 使用正确的总抽样目标计算完成率
        'min_ok': 'N/A',
        'max_ok': 'N/A'
    }
    report.append(total_row)
    
    # 添加各条件统计
    for cond in conditions:
        raw_count = original_data.eval(cond['condition']).sum()
        sampled_count = sampled.eval(cond['condition']).sum()
        
        # 验证min-max限制
        min_ok = sampled_count >= cond['min']
        max_ok = sampled_count <= cond['max'] if cond['max'] != float('inf') else True
        
        row = {
            'condition': cond['condition'],
            'quota': cond['quota'],
            'min': cond['min'],
            'max': cond['max'] if cond['max'] != float('inf') else 'N/A',
            'raw': raw_count,
            'raw_pct': f"{raw_count / len(original_data) * 100:.1f}%" if len(original_data) > 0 else "0.0%",
            'sampled': sampled_count,
            'sampled_pct': f"{sampled_count / len(sampled) * 100:.1f}%" if len(sampled) > 0 else "0.0%",
            'gap': sampled_count - cond['quota'],
            'completion': f"{sampled_count / cond['quota'] * 100:.1f}%" if cond['quota'] > 0 else "N/A",
            'min_ok': '✓' if min_ok else '✗',
            'max_ok': '✓' if max_ok else '✗'
        }
        report.append(row)
    
    return pd.DataFrame(report)

def format_excel(writer, report_df):
    """美化Excel输出格式"""
    workbook = writer.book
    
    # 格式化统计报告工作表
    worksheet = writer.sheets['报告']
    
    # 定义样式
    header_fill = PatternFill(start_color='FFCCCCFF', end_color='FFCCCCFF', fill_type='solid')
    header_font = Font(bold=True, color='FF000000')
    center_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # 应用样式到表头
    for col_num, column in enumerate(report_df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # 调整列宽
    for col_num, column in enumerate(report_df.columns, 1):
        column_width = max(len(str(column)), report_df[column].astype(str).map(len).max())
        worksheet.column_dimensions[chr(64 + col_num)].width = column_width + 4
    
    # 应用样式到数据单元格
    for row_num in range(2, len(report_df) + 2):
        for col_num in range(1, len(report_df.columns) + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
    
    # 高亮显示未达标的配额和min-max限制
    for row_num, row in enumerate(report_df.itertuples(), 2):
        # 处理配额差距
        if hasattr(row, 'gap') and hasattr(row, 'quota') and row.quota > 0:
            gap_cell = worksheet.cell(row=row_num, column=report_df.columns.get_loc('gap') + 1)
            completion_cell = worksheet.cell(row=row_num, column=report_df.columns.get_loc('completion') + 1)
            
            # 如果未达到配额，标黄色（警告）
            if getattr(row, 'gap') < 0:
                gap_cell.fill = PatternFill(start_color='FFFFCC00', end_color='FFFFCC00', fill_type='solid')
                gap_cell.font = Font(bold=True)
                completion_cell.fill = PatternFill(start_color='FFFFCC00', end_color='FFFFCC00', fill_type='solid')
                completion_cell.font = Font(bold=True)
            # 如果达到或超过配额，标绿
            else:
                gap_cell.fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
                completion_cell.fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
        
        # 处理min-max限制
        if hasattr(row, 'min_ok') and row.min_ok != 'N/A':
            min_ok_cell = worksheet.cell(row=row_num, column=report_df.columns.get_loc('min_ok') + 1)
            # 如果不满足最小要求，标红
            if row.min_ok == '✗':
                min_ok_cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                min_ok_cell.font = Font(color='FFFFFFFF', bold=True)
            else:
                min_ok_cell.fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
        
        if hasattr(row, 'max_ok') and row.max_ok != 'N/A':
            max_ok_cell = worksheet.cell(row=row_num, column=report_df.columns.get_loc('max_ok') + 1)
            # 如果超过最大限制，标红
            if row.max_ok == '✗':
                max_ok_cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                max_ok_cell.font = Font(color='FFFFFFFF', bold=True)
            else:
                max_ok_cell.fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')

def main():
    try:
        logging.info("开始执行抽样程序")
        
        # 获取脚本所在目录
        script_dir = os.path.dirname(os.path.abspath(__file__))
        
        # 读取输入文件
        input_file = os.path.join(script_dir, file_name)
        if not os.path.exists(input_file):
            raise FileNotFoundError(f"输入文件 {input_file} 不存在")
            
        logging.info(f"读取输入文件: {input_file}")
        data_df = pd.read_excel(input_file, sheet_name=0)
        cond_df = pd.read_excel(input_file, sheet_name=1)
        
        # 验证输入数据
        validate_input_data(data_df, cond_df)
        
        # 解析抽数条件
        total, conditions = parse_conditions(cond_df)
        
        # 执行线性规划抽样
        original_data = data_df.copy()
        try:
            sampled, full_data_with_sl = linear_programming_sampling(data_df, conditions, total)
            
            # 生成统计报告
            report_df = generate_report(sampled, conditions, original_data, total)
            
            # 保存结果
            output_file = os.path.join(script_dir, 'output.xlsx')
            logging.info(f"保存结果到: {output_file}")
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # 输出完整数据集（包含SL标记列）
                full_data_with_sl.to_excel(writer, sheet_name='样本', index=False)
                report_df.to_excel(writer, sheet_name='报告', index=False)
                format_excel(writer, report_df)
            
            logging.info(f"抽样成功，共抽取 {len(sampled)} 个样本，详细报告已保存到 {output_file}")
            print(f"\n抽样成功，共抽取 {len(sampled)} 个样本，详细报告已保存到 {output_file}")
            print(f"日志文件保存在: {log_file}")
            
        except ValueError as ve:
            # 抽样失败的特定处理
            logging.error(f"抽样失败: {str(ve)}")
            print(f"\n抽样失败: {str(ve)}")
            print("请调整条件配置或数据后重试")
            print(f"详细错误信息请查看日志: {log_file}")
        
    except Exception as e:
        logging.error(f"程序执行出错: {str(e)}", exc_info=True)
        print(f"\n错误: {str(e)}")
        print(f"详细错误信息请查看日志: {log_file}")

if __name__ == '__main__':
    main()